
# Copyright 2010-2019 Amazon.com, Inc. or its affiliates. All Rights Reserved.
#
# This file is licensed under the Apache License, Version 2.0 (the "License").
# You may not use this file except in compliance with the License. A copy of the
# License is located at
#
# http://aws.amazon.com/apache2.0/
#
# This file is distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS
# OF ANY KIND, either express or implied. See the License for the specific
# language governing permissions and limitations under the License.

# import module
import openpyxl
import boto3
translate = boto3.client('translate')

# load excel with its path
wrkbk = openpyxl.load_workbook("input.xlsx")

sh = wrkbk.active

# iterate through excel and display data
for i in range(1, sh.max_row+1):
	toTranslate=sh.cell(row=i, column=2).value
	result = translate.translate_text(Text=toTranslate,
											SourceLanguageCode="hu",
											TargetLanguageCode="pt")
	translated=result["TranslatedText"]
	print(translated)
	sh.cell(row=i,column=4).value = translated

wrkbk.save('output.xlsx')
